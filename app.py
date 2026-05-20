import streamlit as st
import io, os, json, smtplib, copy
from datetime import datetime, date
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.header import Header
from email import encoders
import openpyxl
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from pypdf import PdfReader, PdfWriter
from reportlab.lib.utils import simpleSplit
# ── สำหรับเข้ารหัส Excel (ต้อง pip install msoffcrypto-tool) ───────────────────
try:
    import msoffcrypto
    HAS_MSOFFCRYPTO = True
except ImportError:
    HAS_MSOFFCRYPTO = False
# 🔐 รหัสผ่าน Excel — แก้ตรงนี้
EXCEL_PASSWORD = "cpram2024"
# 🔑 รหัสผ่านสำหรับเข้าหน้า "ผู้ดูแลระบบ" — แก้ตรงนี้
ADMIN_PASSWORD = "admin123"
# ══════════════════════════════════════════════════════════════════════════════
# ⚙️  SMTP CONFIG — แก้ตรงนี้ครั้งเดียว แล้วผู้ใช้ไม่ต้องกรอกอีก
# ══════════════════════════════════════════════════════════════════════════════
DEFAULT_SMTP = {
    "host":     "smtp.gmail.com",
    "port":     465,
    "user":     "natchathompad@gmail.com",
    "password": "ngxzkqhlffghvvpv",
}
def get_smtp_config():
    try:
        return {
            "host":     st.secrets["smtp"]["host"],
            "port":     int(st.secrets["smtp"]["port"]),
            "user":     st.secrets["smtp"]["user"],
            "password": st.secrets["smtp"]["password"],
        }
    except (KeyError, FileNotFoundError, Exception):
        return DEFAULT_SMTP
# ─── page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="แบบสอบถาม CPRAM – ผักสลัด", layout="wide")
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template.pdf")
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data_cpram.xlsx")
# ─── Thai font setup ────────────────────────────────────────────────────────────
THAI_FONT = None
_local = os.path.join(os.path.dirname(__file__), "fonts", "THSarabunNew.ttf")
if os.path.exists(_local):
    pdfmetrics.registerFont(TTFont("ThaiFont", _local))
    THAI_FONT = "ThaiFont"
else:
    import glob
    for pattern in ["C:/Windows/Fonts/**/*.ttf", "/usr/share/fonts/**/*.ttf"]:
        for h in glob.glob(pattern, recursive=True):
            if any(k in h.lower() for k in ["sarabun","thai","garuda","cordia"]):
                try:
                    pdfmetrics.registerFont(TTFont("ThaiFont", h))
                    THAI_FONT = "ThaiFont"
                    break
                except: continue
        if THAI_FONT: break
FONT_NAME = THAI_FONT if THAI_FONT else "Helvetica"
# ─── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main { font-family: 'Sarabun', sans-serif; }
    .section-header {
        color: #2e7d32;
        font-size: 1.15rem;
        font-weight: 700;
        border-bottom: 2px solid #2e7d32;
        padding-bottom: 6px;
        margin-bottom: 16px;
        margin-top: 8px;
    }
    .item-card {
        background: #f1f8e9;
        border: 1px solid #a5d6a7;
        border-radius: 10px;
        padding: 16px 20px 8px;
        margin-bottom: 18px;
    }
    .item-title {
        font-weight: 700;
        font-size: 1rem;
        color: #1b5e20;
        margin-bottom: 12px;
    }
    .stButton>button {
        background: #2e7d32;
        color: white;
        border-radius: 8px;
        border: none;
        padding: 0.5rem 1.5rem;
        font-weight: 600;
    }
    .stButton>button:hover { background: #1b5e20; }
    .delete-btn>button {
        background: #c62828 !important;
        font-size: 0.85rem;
        padding: 0.3rem 0.9rem;
    }
    .add-btn>button {
        background: white !important;
        color: #2e7d32 !important;
        border: 2px dashed #2e7d32 !important;
        width: 100%;
    }
</style>
""", unsafe_allow_html=True)
# ─── session state ──────────────────────────────────────────────────────────────
if "item_list" not in st.session_state:
    st.session_state.item_list = [{}]
if "preview_pdf" not in st.session_state:
    st.session_state.preview_pdf = None
if "preview_data" not in st.session_state:
    st.session_state.preview_data = None
if "confirm_clear" not in st.session_state:
    st.session_state.confirm_clear = False
if "confirm_delete_file" not in st.session_state:
    st.session_state.confirm_delete_file = False
def add_item():
    st.session_state.item_list.append({})
def remove_item(idx):
    if len(st.session_state.item_list) > 1:
        st.session_state.item_list.pop(idx)
def clear_preview():
    st.session_state.preview_pdf = None
    st.session_state.preview_data = None
# ─── Thai Data with Zipcode from JSON ──────────────────────────────────────────
@st.cache_data
def load_geo_data():
    file_path = os.path.join(os.path.dirname(__file__), "thailand_geo_with_zip.json")
    with open(file_path, encoding="utf-8") as f:
        return json.load(f)
GEO_DATA = load_geo_data()
def get_provinces(country):
    if country == "ไทย" and GEO_DATA:
        return sorted(list(GEO_DATA.keys()))
    return []
def get_districts(country, province):
    if country == "ไทย" and GEO_DATA and province in GEO_DATA:
        return sorted(list(GEO_DATA[province].keys()))
    return []
def get_subdistricts(country, province, district):
    if country == "ไทย" and GEO_DATA and province in GEO_DATA and district in GEO_DATA[province]:
        return sorted(list(GEO_DATA[province][district].keys()))
    return []
def get_zipcode(province, district, subdistrict):
    try:
        return GEO_DATA[province][district][subdistrict]
    except KeyError:
        return ""
# ─── EXCEL helper ───────────────────────────────────────────────────────────────
def save_to_excel(form_data: dict):
    headers = [
        "วันที่บันทึก", "ผู้ส่งมอบ", "วันที่ส่งวัตถุดิบ", "เวลาส่ง",
        "อีเมล", "ลงชื่อผู้กรอก", "ประเทศแหล่งปลูก(default)",
        "ชนิดวัตถุดิบ", "Code", "จำนวน(KG)",
        "วันที่เก็บเกี่ยว", "เวลาเก็บเกี่ยว", "วันที่ล้างความสะอาด",
        "เวลาล้างความสะอาด", "ชื่อผู้ปลูก", "เลขที่GAP",
        "รหัสไร่", "ที่อยู่เลขที่", "หมู่ที่",
        "ประเทศ", "จังหวัด", "อำเภอ/เมือง", "ตำบล/เขต",
        "รหัสไปรษณีย์", "สายพันธุ์", "ลักษณะการปลูก", "ลักษณะสถานที่ปลูก",
    ]
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CPRAM Data"
        ws.append(headers)
        wb.save(EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    for item in form_data.get("items", []):
        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            form_data.get("supplier", ""),
            form_data.get("delivery_date", ""),
            form_data.get("delivery_time", ""),
            form_data.get("email", ""),
            form_data.get("signer", ""),
            form_data.get("default_country", ""),
            item.get("material_type", ""),
            item.get("code", ""),
            item.get("quantity", ""),
            item.get("harvest_date", ""),
            item.get("harvest_time", ""),
            item.get("clean_date", ""),
            item.get("clean_time", ""),
            item.get("grower_name", ""),
            item.get("gap_number", ""),
            item.get("farm_code", ""),
            item.get("address_no", ""),
            item.get("moo", ""),
            item.get("country", ""),
            item.get("province", ""),
            item.get("district", ""),
            item.get("subdistrict", ""),
            item.get("postal_code", ""),
            item.get("variety", ""),
            item.get("growing_type", ""),
            item.get("growing_condition", ""),
        ]
        ws.append(row)
    wb.save(EXCEL_PATH)
# ─── PDF generation ─────────────────────────────────────────────────────────────
def draw_text(c, text, x, y, font_size=9, max_width=None):
    if not text:
        return
    c.setFont(FONT_NAME, font_size)
    if max_width:
        lines = simpleSplit(str(text), FONT_NAME, font_size, max_width)
        for i, line in enumerate(lines):
            c.drawString(x, y - i * (font_size + 2), line)
    else:
        c.drawString(x, y, str(text))

def _draw_item_overlay(c, item, supplier, delivery_date, signer, fs=14):
    """วาดข้อมูลรายการวัตถุดิบหนึ่งชุดบน canvas overlay (ใช้ template หน้าแรก)"""
    # Header
    draw_text(c, supplier,      220, 728, fs, max_width=230)
    draw_text(c, delivery_date, 465, 728, fs, max_width=100)
    # ข้อมูลวัตถุดิบ
    draw_text(c, item.get("material_type",""), 260, 696, fs, max_width=120)
    draw_text(c, item.get("code",""),          390, 696, fs, max_width=80)
    draw_text(c, item.get("quantity",""),      490, 696, fs, max_width=60)
    draw_text(c, item.get("variety",""),           145, 667, fs, max_width=130)
    draw_text(c, item.get("growing_type",""),      310, 667, fs, max_width=130)
    draw_text(c, item.get("growing_condition",""), 470, 667, fs, max_width=90)
    draw_text(c, item.get("harvest_date",""), 210, 638, fs, max_width=150)
    draw_text(c, item.get("harvest_time",""), 420, 638, fs, max_width=130)
    draw_text(c, item.get("clean_date",""), 190, 609, fs, max_width=150)
    draw_text(c, item.get("clean_time",""), 420, 609, fs, max_width=130)
    draw_text(c, item.get("grower_name",""), 160, 580, fs, max_width=330)
    draw_text(c, item.get("gap_number",""), 155, 551, fs, max_width=200)
    draw_text(c, item.get("farm_code",""),  370, 551, fs, max_width=120)
    addr = " ".join(filter(None, [
        item.get("address_no",""),
        f"หมู่ {item.get('moo','')}" if item.get("moo","") else "",
        item.get("subdistrict",""),
        item.get("district",""),
        item.get("province",""),
        item.get("postal_code",""),
    ]))
    draw_text(c, addr, 115, 522, fs, max_width=440)
    # Footer signature
    draw_text(c, signer,        420, 62, fs)
    draw_text(c, delivery_date, 420, 42, fs)

def generate_pdf(form_data: dict) -> bytes:
    """
    สร้าง PDF โดย *วน loop สร้างหน้าใหม่ต่อ 1 รายการวัตถุดิบ*
    ใช้ template.pdf หน้าแรกเป็นแม่แบบของทุกหน้ารายการ
    """
    reader = PdfReader(TEMPLATE_PATH)
    writer = PdfWriter()
    supplier      = form_data.get("supplier", "")
    delivery_date = form_data.get("delivery_date", "")
    signer        = form_data.get("signer", "")
    items         = form_data.get("items", []) or [{}]
    W, H = 595.32, 841.92
    fs = 14
    # ── สร้าง 1 หน้า ต่อ 1 รายการวัตถุดิบ ─────────────────────────────────────
    for item in items:
        overlay_buf = io.BytesIO()
        c = canvas.Canvas(overlay_buf, pagesize=(W, H))
        c.setFillColorRGB(0, 0, 0)
        _draw_item_overlay(c, item, supplier, delivery_date, signer, fs)
        c.save()
        overlay_buf.seek(0)
        overlay_reader = PdfReader(overlay_buf)
        # ใช้หน้าแรกของ template เป็นแม่แบบของทุกหน้ารายการ
        page = copy.copy(reader.pages[0])
        page.merge_page(overlay_reader.pages[0])
        writer.add_page(page)
    # ── หน้าที่เหลือของ template (ถ้ามี) เติมท้ายแค่ครั้งเดียว ────────────────
    for page in reader.pages[1:]:
        extra_buf = io.BytesIO()
        ec = canvas.Canvas(extra_buf, pagesize=(W, H))
        ec.setFillColorRGB(0, 0, 0)
        draw_text(ec, signer,        420, 62, fs)
        draw_text(ec, delivery_date, 420, 42, fs)
        ec.save()
        extra_buf.seek(0)
        extra_reader = PdfReader(extra_buf)
        p = copy.copy(page)
        p.merge_page(extra_reader.pages[0])
        writer.add_page(p)
    out_buf = io.BytesIO()
    writer.write(out_buf)
    out_buf.seek(0)
    return out_buf.read()
# ─── Email send ─────────────────────────────────────────────────────────────────
def send_email(to_addr: str, pdf_bytes: bytes, smtp_cfg: dict) -> tuple[bool, str]:
    try:
        msg = MIMEMultipart()
        msg["From"]    = smtp_cfg["user"]
        msg["To"]      = to_addr
        msg["Subject"] = Header("แบบสอบถามประจำวัน CPRAM – ผักสลัด", "utf-8")
        msg.attach(MIMEText("กรุณาดูเอกสารแบบสอบถามที่แนบมาด้วยครับ/ค่ะ", "plain", "utf-8"))
        part = MIMEBase("application", "octet-stream")
        part.set_payload(pdf_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", 'attachment; filename="cpram_questionnaire.pdf"')
        msg.attach(part)
        port = int(smtp_cfg["port"])
        if port == 465:
            server = smtplib.SMTP_SSL(smtp_cfg["host"], port, timeout=30)
        else:
            server = smtplib.SMTP(smtp_cfg["host"], port, timeout=30)
            server.starttls()
        server.login(smtp_cfg["user"], smtp_cfg["password"])
        server.send_message(msg)
        server.quit()
        return True, ""
    except smtplib.SMTPAuthenticationError as e:
        return False, f"รหัสผ่านไม่ถูกต้อง (ต้องใช้ App Password 16 หลัก): {e}"
    except Exception as e:
        return False, str(e)
# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("<h2 style='color:#1b5e20;text-align:center;'>📋 แบบสอบถามประจำวัน CPRAM – ผักสลัด</h2>", unsafe_allow_html=True)
# ── ส่วนที่ 1 ──────────────────────────────────────────────────────────────────
st.markdown('<div class="section-header">ส่วนที่ 1 — ข้อมูลผู้ส่งมอบและการส่งมอบ</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns([2, 2, 2])
with col1:
    supplier = st.text_input("ผู้ส่งมอบ (Supplier) *", key="supplier")
with col2:
    delivery_date = st.date_input("วันที่ส่งวัตถุดิบ *", key="delivery_date", value=None)
with col3:
    delivery_time = st.text_input("เวลาส่ง (เช่น 08.00 - 09.00)", key="delivery_time", placeholder="08.00 - 09.00")
col4, col5, col6 = st.columns([2, 2, 2])
with col4:
    email = st.text_input("อีเมลของผู้ส่งมอบ (สำหรับรับ PDF) *", key="email")
with col5:
    signer = st.text_input("ลงชื่อผู้กรอก", key="signer")
with col6:
    default_country = st.selectbox("ประเทศแหล่งปลูก (default)", ["ประเทศไทย", "จีน", "ญี่ปุ่น", "เกาหลี", "อื่นๆ"], key="default_country")
st.markdown("---")
# ── ส่วนที่ 2 ──────────────────────────────────────────────────────────────────
st.markdown('<div class="section-header">ส่วนที่ 2 — รายการวัตถุดิบ <span style="font-size:0.85rem;font-weight:400;">(เพิ่มได้ไม่จำกัด)</span></div>', unsafe_allow_html=True)
GROWING_TYPES = ["- เลือก -", "ปลูกดินยกพื้น", "ปลูกดินไม่ยกพื้น", "ปลูกไฮโดรโปนิกส์"]
GROWING_CONDITIONS = ["- เลือก -", "ระบบเปิด", "ระบบปิด"]
items_data = []
for i, _ in enumerate(st.session_state.item_list):
    st.markdown(f'<div class="item-card"><div class="item-title">รายการที่ {i+1}</div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns([3, 2, 2])
    with c1:
        mat = st.text_input("ชนิดวัตถุดิบที่ส่งให้ทาง CPRAM *", key=f"mat_{i}")
    with c2:
        code = st.text_input("Code (เช่น 71000277)", key=f"code_{i}")
    with c3:
        qty = st.text_input("จำนวน (KG) *", key=f"qty_{i}")
    c4, c5, c6 = st.columns([2, 2, 2])
    with c4:
        h_date = st.date_input("วันที่เก็บเกี่ยว", key=f"hdate_{i}", value=None)
    with c5:
        h_time = st.text_input("เวลาเก็บเกี่ยว (เช่น 08.00 - 09.00)", key=f"htime_{i}", placeholder="08.00 - 09.00")
    with c6:
        c_date = st.date_input("วันที่ล้างทำความสะอาด", key=f"cdate_{i}", value=None)
    c7, c8, c9 = st.columns([2, 2, 2])
    with c7:
        c_time = st.text_input("เวลาล้างทำความสะอาด (เช่น 08.00 - 09.00)", key=f"ctime_{i}", placeholder="08.00 - 09.00")
    with c8:
        grower = st.text_input("ชื่อผู้ปลูก", key=f"grower_{i}")
    with c9:
        gap = st.text_input("เลขที่ GAP", key=f"gap_{i}")
    c10, c11, c12 = st.columns([2, 2, 2])
    with c10:
        farm_code = st.text_input("รหัสไร่", key=f"fcode_{i}")
    with c11:
        addr_no = st.text_input("ที่อยู่เลขที่", key=f"addr_{i}")
    with c12:
        moo = st.text_input("หมู่ที่", key=f"moo_{i}")
    st.markdown("📍 **ที่อยู่แหล่งปลูก**")
    cc1, cc2, cc3, cc4 = st.columns(4)
    with cc1:
        country_opts = ["ไทย"]
        country = st.selectbox("ประเทศ", country_opts, key=f"cntry_{i}")
    with cc2:
        prov_opts = ["- เลือก -"] + get_provinces(country)
        province = st.selectbox("จังหวัด/มณฑล", prov_opts, key=f"prov_{i}")
    with cc3:
        dist_opts = ["- เลือกจังหวัดก่อน -"]
        if province != "- เลือก -":
            dist_opts = ["- เลือก -"] + get_districts(country, province)
        district = st.selectbox("อำเภอ/เมือง", dist_opts, key=f"dist_{i}")
    with cc4:
        sub_opts = ["- เลือกอำเภอก่อน -"]
        if district not in ("- เลือก -", "- เลือกจังหวัดก่อน -"):
            subs = get_subdistricts(country, province, district)
            sub_opts = ["- เลือก -"] + subs
        subdistrict = st.selectbox("ตำบล/เขต", sub_opts, key=f"sub_{i}")
    auto_zip = ""
    if country == "ไทย" and subdistrict not in ("- เลือก -", "- เลือกอำเภอก่อน -"):
        auto_zip = get_zipcode(province, district, subdistrict)
    cc5, cc6, cc7, cc8 = st.columns(4)
    with cc5:
        st.text_input("รหัสไปรษณีย์", value=auto_zip, key=f"postal_display_{i}", disabled=True)
        postal = auto_zip
    with cc6:
        variety = st.text_input("สายพันธุ์", key=f"variety_{i}")
    with cc7:
        g_type = st.selectbox("ลักษณะการปลูก", GROWING_TYPES, key=f"gtype_{i}")
    with cc8:
        g_cond = st.selectbox("ลักษณะสถานที่ปลูก", GROWING_CONDITIONS, key=f"gcond_{i}")
    if len(st.session_state.item_list) > 1:
        _, del_col = st.columns([6, 1])
        with del_col:
            st.markdown('<div class="delete-btn">', unsafe_allow_html=True)
            if st.button("✕ ลบรายการนี้", key=f"del_{i}"):
                remove_item(i)
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    items_data.append({
        "material_type": mat,
        "code": code,
        "quantity": qty,
        "harvest_date": h_date.strftime("%d/%m/%Y") if h_date else "",
        "harvest_time": h_time,
        "clean_date": c_date.strftime("%d/%m/%Y") if c_date else "",
        "clean_time": c_time,
        "grower_name": grower,
        "gap_number": gap,
        "farm_code": farm_code,
        "address_no": addr_no,
        "moo": moo,
        "country": country,
        "province": province if province != "- เลือก -" else "",
        "district": district if district not in ("- เลือก -","- เลือกจังหวัดก่อน -") else "",
        "subdistrict": subdistrict if subdistrict not in ("- เลือก -","- เลือกอำเภอก่อน -") else "",
        "postal_code": postal,
        "variety": variety,
        "growing_type": g_type if g_type != "- เลือก -" else "",
        "growing_condition": g_cond if g_cond != "- เลือก -" else "",
    })
st.markdown('<div class="add-btn">', unsafe_allow_html=True)
if st.button("＋ เพิ่มรายการวัตถุดิบ"):
    add_item()
    st.rerun()
st.markdown('</div>', unsafe_allow_html=True)
st.markdown("---")
# ══════════════════════════════════════════════════════════════════════════════
# 🚀 ขั้นที่ 1 — ตรวจสอบและสร้างตัวอย่าง PDF
# ══════════════════════════════════════════════════════════════════════════════
if st.button("👁️ ดูตัวอย่างก่อนส่ง", use_container_width=True, type="primary"):
    errors = []
    if not supplier:       errors.append("กรุณากรอก ผู้ส่งมอบ")
    if not delivery_date:  errors.append("กรุณาเลือก วันที่ส่งวัตถุดิบ")
    if not delivery_time:  errors.append("กรุณากรอก เวลาส่ง")
    if not email:          errors.append("กรุณากรอก อีเมล")
    if not signer:         errors.append("กรุณากรอก ลงชื่อผู้กรอก")
    for idx, it in enumerate(items_data):
        n = idx + 1
        if not it["material_type"]:     errors.append(f"รายการที่ {n}: กรุณากรอก ชนิดวัตถุดิบ")
        if not it["code"]:              errors.append(f"รายการที่ {n}: กรุณากรอก Code")
        if not it["quantity"]:          errors.append(f"รายการที่ {n}: กรุณากรอก จำนวน")
        if not it["harvest_date"]:      errors.append(f"รายการที่ {n}: กรุณาเลือก วันที่เก็บเกี่ยว")
        if not it["harvest_time"]:      errors.append(f"รายการที่ {n}: กรุณากรอก เวลาเก็บเกี่ยว")
        if not it["clean_date"]:        errors.append(f"รายการที่ {n}: กรุณาเลือก วันที่ล้าง")
        if not it["clean_time"]:        errors.append(f"รายการที่ {n}: กรุณากรอก เวลาล้าง")
        if not it["grower_name"]:       errors.append(f"รายการที่ {n}: กรุณากรอก ชื่อผู้ปลูก")
        if not it["gap_number"]:        errors.append(f"รายการที่ {n}: กรุณากรอก เลขที่ GAP")
        if not it["farm_code"]:         errors.append(f"รายการที่ {n}: กรุณากรอก รหัสไร่")
        if not it["address_no"]:        errors.append(f"รายการที่ {n}: กรุณากรอก ที่อยู่เลขที่")
        if not it["province"]:          errors.append(f"รายการที่ {n}: กรุณาเลือก จังหวัด")
        if not it["district"]:          errors.append(f"รายการที่ {n}: กรุณาเลือก อำเภอ")
        if not it["subdistrict"]:       errors.append(f"รายการที่ {n}: กรุณาเลือก ตำบล")
        if not it["variety"]:           errors.append(f"รายการที่ {n}: กรุณากรอก สายพันธุ์")
        if not it["growing_type"]:      errors.append(f"รายการที่ {n}: กรุณาเลือก ลักษณะการปลูก")
        if not it["growing_condition"]: errors.append(f"รายการที่ {n}: กรุณาเลือก ลักษณะสถานที่ปลูก")
    if errors:
        for e in errors:
            st.error(e)
    else:
        form_data = {
            "supplier":        supplier,
            "delivery_date":   delivery_date.strftime("%d/%m/%Y") if delivery_date else "",
            "delivery_time":   delivery_time,
            "email":           email,
            "signer":          signer,
            "default_country": default_country,
            "items":           items_data,
        }
        with st.spinner("กำลังสร้างตัวอย่าง PDF..."):
            try:
                pdf_bytes = generate_pdf(form_data)
                st.session_state.preview_pdf  = pdf_bytes
                st.session_state.preview_data = form_data
                st.rerun()
            except Exception as ex:
                st.error(f"สร้าง PDF ไม่สำเร็จ: {ex}")

# ══════════════════════════════════════════════════════════════════════════════
# 🔍 ขั้นที่ 2 — แสดงตัวอย่าง + ปุ่มยืนยัน/แก้ไข
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.preview_pdf and st.session_state.preview_data:
    st.markdown("---")
    st.markdown('<div class="section-header">🔍 ตรวจสอบข้อมูลก่อนส่ง</div>', unsafe_allow_html=True)
    pd_preview = st.session_state.preview_data
    # สรุปหัวข้อ
    st.info(
        f"**ผู้ส่งมอบ:** {pd_preview['supplier']}  \n"
        f"**วันที่/เวลาส่ง:** {pd_preview['delivery_date']}  เวลา {pd_preview['delivery_time']}  \n"
        f"**ผู้กรอก:** {pd_preview['signer']}  \n"
        f"**ส่งไปที่อีเมล:** {pd_preview['email']}  \n"
        f"**จำนวนรายการวัตถุดิบ:** {len(pd_preview['items'])} รายการ"
    )
    # สรุปรายการ
    with st.expander(f"📋 ดูรายการวัตถุดิบทั้งหมด ({len(pd_preview['items'])} รายการ)", expanded=True):
        for i, it in enumerate(pd_preview["items"], 1):
            st.markdown(
                f"**รายการที่ {i}:** {it['material_type']} | Code: {it['code']} | {it['quantity']} KG | สายพันธุ์: {it['variety']}  \n"
                f"&nbsp;&nbsp;&nbsp;&nbsp;เก็บเกี่ยว: {it['harvest_date']} {it['harvest_time']} | ล้าง: {it['clean_date']} {it['clean_time']}  \n"
                f"&nbsp;&nbsp;&nbsp;&nbsp;ผู้ปลูก: {it['grower_name']} | GAP: {it['gap_number']} | รหัสไร่: {it['farm_code']}  \n"
                f"&nbsp;&nbsp;&nbsp;&nbsp;ที่อยู่: {it['address_no']} หมู่ {it['moo']} ต.{it['subdistrict']} อ.{it['district']} จ.{it['province']} {it['postal_code']}"
            )
            st.markdown("---")
    # ปุ่มดาวน์โหลด PDF ดูตัวอย่างก่อน
    st.download_button(
        label="📥 ดาวน์โหลด PDF เพื่อดูตัวอย่าง",
        data=st.session_state.preview_pdf,
        file_name=f"preview_cpram_{pd_preview['supplier']}.pdf",
        mime="application/pdf",
        use_container_width=True,
    )
    # ปุ่มยืนยัน / กลับไปแก้ไข
    cf1, cf2 = st.columns(2)
    with cf1:
        if st.button("✅ ยืนยันและส่งอีเมล", use_container_width=True, type="primary", key="confirm_send"):
            with st.spinner("กำลังส่งอีเมล..."):
                # 1) บันทึก Excel
                try:
                    save_to_excel(pd_preview)
                except Exception as ex:
                    st.warning(f"บันทึก Excel ไม่สำเร็จ: {ex}")
                # 2) ส่งอีเมล
                smtp_cfg = get_smtp_config()
                sent, err = send_email(pd_preview["email"], st.session_state.preview_pdf, smtp_cfg)
                if sent:
                    st.success(f"✅ ส่ง PDF ไปยัง **{pd_preview['email']}** เรียบร้อยแล้ว!")
                    st.balloons()
                    clear_preview()
                else:
                    st.error(f"❌ ส่งอีเมลไม่สำเร็จ: {err}")
                    st.info("กรุณาดาวน์โหลด PDF ด้านบน แล้วแจ้งผู้ดูแลระบบ")
    with cf2:
        if st.button("✏️ กลับไปแก้ไขข้อมูล", use_container_width=True, key="cancel_preview"):
            clear_preview()
            st.rerun()
# ── ส่วนสำหรับผู้ดูแลระบบ ──────────────────────────────────────────────────────
def get_encrypted_excel_bytes() -> bytes:
    with open(EXCEL_PATH, "rb") as f:
        plain_bytes = f.read()
    if not HAS_MSOFFCRYPTO:
        return plain_bytes
    try:
        plain_buf = io.BytesIO(plain_bytes)
        enc_buf = io.BytesIO()
        office = msoffcrypto.OfficeFile(plain_buf)
        office.encrypt(EXCEL_PASSWORD, enc_buf)
        enc_buf.seek(0)
        return enc_buf.read()
    except Exception as ex:
        st.warning(f"เข้ารหัสไม่สำเร็จ ({ex}) — ดาวน์โหลดเป็นไฟล์ปกติแทน")
        return plain_bytes
def clear_excel_data():
    """ลบข้อมูลใน Excel แต่เก็บหัวคอลัมน์ไว้"""
    if not os.path.exists(EXCEL_PATH):
        return False, "ยังไม่มีไฟล์ Excel"
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)   # ลบทุกแถวยกเว้นหัวคอลัมน์
        wb.save(EXCEL_PATH)
        return True, ""
    except Exception as ex:
        return False, str(ex)

def delete_excel_file():
    """ลบไฟล์ Excel ทั้งหมด"""
    if not os.path.exists(EXCEL_PATH):
        return False, "ยังไม่มีไฟล์ Excel"
    try:
        os.remove(EXCEL_PATH)
        return True, ""
    except Exception as ex:
        return False, str(ex)

# ══════════════════════════════════════════════════════════════════════════════
# 📊 Dashboard / 🔍 Traceability helpers
# ══════════════════════════════════════════════════════════════════════════════
def _to_num(s):
    """แปลงคอลัมน์เป็นตัวเลข (errors='coerce' = แปลงไม่ได้ให้เป็น NaN)"""
    return pd.to_numeric(s, errors="coerce")

def _to_date(s):
    """แปลงคอลัมน์เป็น date (รองรับทั้ง dd/mm/yyyy และ yyyy-mm-dd hh:mm:ss)"""
    return pd.to_datetime(s, errors="coerce", dayfirst=True)

def render_dashboard(df: pd.DataFrame):
    """แสดง Dashboard — ภาพรวม + กราฟ + ตารางล่าสุด"""
    if df.empty:
        st.info("ยังไม่มีข้อมูล")
        return
    # ── ตัวกรองช่วงเวลา ─────────────────────────────────────────────
    df["_recorded_dt"] = _to_date(df["วันที่บันทึก"])
    df["_delivery_dt"] = _to_date(df["วันที่ส่งวัตถุดิบ"])
    df["_qty_num"]     = _to_num(df["จำนวน(KG)"])
    min_d = df["_recorded_dt"].min()
    max_d = df["_recorded_dt"].max()
    fc1, fc2 = st.columns(2)
    with fc1:
        d_from = st.date_input("จากวันที่", value=min_d.date() if pd.notna(min_d) else None, key="dash_from")
    with fc2:
        d_to = st.date_input("ถึงวันที่", value=max_d.date() if pd.notna(max_d) else None, key="dash_to")
    # apply filter
    fdf = df.copy()
    if d_from:
        fdf = fdf[fdf["_recorded_dt"] >= pd.Timestamp(d_from)]
    if d_to:
        fdf = fdf[fdf["_recorded_dt"] <= pd.Timestamp(d_to) + pd.Timedelta(days=1)]
    # ── KPI ──────────────────────────────────────────────────────
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("📦 จำนวนรายการ", f"{len(fdf):,}")
    m2.metric("⚖️ น้ำหนักรวม (KG)", f"{fdf['_qty_num'].sum():,.0f}")
    m3.metric("🏭 ผู้ส่งมอบ (ราย)", f"{fdf['ผู้ส่งมอบ'].nunique():,}")
    m4.metric("🥬 ชนิดวัตถุดิบ", f"{fdf['ชนิดวัตถุดิบ'].nunique():,}")
    if fdf.empty:
        st.warning("ไม่มีข้อมูลในช่วงเวลาที่เลือก")
        return
    # ── กราฟรายวัน ───────────────────────────────────────────────
    st.markdown("#### 📈 น้ำหนักวัตถุดิบที่ส่งรายวัน (KG)")
    daily = (fdf.dropna(subset=["_delivery_dt"])
                 .groupby(fdf["_delivery_dt"].dt.date)["_qty_num"].sum())
    if not daily.empty:
        st.bar_chart(daily)
    else:
        st.caption("ไม่มีข้อมูลวันส่งวัตถุดิบ")
    # ── Top suppliers / materials ────────────────────────────────
    g1, g2 = st.columns(2)
    with g1:
        st.markdown("#### 🏆 Top 10 ผู้ส่งมอบ (ตามน้ำหนัก)")
        top_sup = fdf.groupby("ผู้ส่งมอบ")["_qty_num"].sum().sort_values(ascending=False).head(10)
        if not top_sup.empty:
            st.bar_chart(top_sup)
    with g2:
        st.markdown("#### 🥬 Top 10 ชนิดวัตถุดิบ (ตามน้ำหนัก)")
        top_mat = fdf.groupby("ชนิดวัตถุดิบ")["_qty_num"].sum().sort_values(ascending=False).head(10)
        if not top_mat.empty:
            st.bar_chart(top_mat)
    g3, g4 = st.columns(2)
    with g3:
        st.markdown("#### 🌾 Top 10 ผู้ปลูก")
        top_grower = fdf.groupby("ชื่อผู้ปลูก").size().sort_values(ascending=False).head(10)
        if not top_grower.empty:
            st.bar_chart(top_grower)
    with g4:
        st.markdown("#### 📍 Top 10 จังหวัดแหล่งปลูก")
        top_prov = fdf.groupby("จังหวัด").size().sort_values(ascending=False).head(10)
        if not top_prov.empty:
            st.bar_chart(top_prov)
    # ── ส่งล่าสุด ────────────────────────────────────────────────
    st.markdown("#### 🕐 รายการที่ส่งล่าสุด 10 รายการ")
    recent_cols = ["วันที่บันทึก","ผู้ส่งมอบ","ชนิดวัตถุดิบ","Code","จำนวน(KG)","ชื่อผู้ปลูก","จังหวัด"]
    recent_cols = [c for c in recent_cols if c in fdf.columns]
    st.dataframe(
        fdf.sort_values("_recorded_dt", ascending=False).head(10)[recent_cols],
        use_container_width=True, hide_index=True
    )

def render_traceability(df: pd.DataFrame):
    """ระบบสืบย้อนวัตถุดิบ — ค้นหาแล้วแสดง timeline ของแต่ละรายการ"""
    if df.empty:
        st.info("ยังไม่มีข้อมูล")
        return
    st.caption("🔎 สืบย้อนวัตถุดิบจากต้นทางถึงการส่งมอบ — ค้นหาด้วย Code, GAP, รหัสไร่, ผู้ปลูก, ผู้ส่งมอบ ฯลฯ")
    sc1, sc2 = st.columns([1, 2])
    with sc1:
        field_map = {
            "Code":            "Code",
            "เลขที่ GAP":     "เลขที่GAP",
            "รหัสไร่":         "รหัสไร่",
            "ชื่อผู้ปลูก":     "ชื่อผู้ปลูก",
            "ผู้ส่งมอบ":       "ผู้ส่งมอบ",
            "ชนิดวัตถุดิบ":   "ชนิดวัตถุดิบ",
            "จังหวัด":         "จังหวัด",
        }
        search_field = st.selectbox("ค้นหาในคอลัมน์", list(field_map.keys()), key="trace_field")
    with sc2:
        search_term = st.text_input("คำค้น (ค้นบางส่วนของคำได้)", key="trace_term")
    # ตัวกรองวันที่ส่ง
    fc1, fc2 = st.columns(2)
    with fc1:
        date_from = st.date_input("วันที่ส่งตั้งแต่", value=None, key="trace_from")
    with fc2:
        date_to = st.date_input("ถึง", value=None, key="trace_to")
    # apply
    result = df.copy()
    col = field_map[search_field]
    if search_term and col in result.columns:
        result = result[result[col].astype(str).str.contains(search_term, case=False, na=False)]
    result["_delivery_dt"] = _to_date(result["วันที่ส่งวัตถุดิบ"])
    if date_from:
        result = result[result["_delivery_dt"] >= pd.Timestamp(date_from)]
    if date_to:
        result = result[result["_delivery_dt"] <= pd.Timestamp(date_to) + pd.Timedelta(days=1)]
    st.success(f"พบ **{len(result)}** รายการที่ตรงกับเงื่อนไข")
    if result.empty:
        return
    # แสดงผลเป็น expander หนึ่งอันต่อหนึ่ง record
    for idx, row in result.head(50).iterrows():
        title = (
            f"📦 {row.get('ชนิดวัตถุดิบ','')} "
            f"({row.get('Code','')}) — {row.get('จำนวน(KG)','')} KG  |  "
            f"ส่งโดย {row.get('ผู้ส่งมอบ','')}  |  วันส่ง {row.get('วันที่ส่งวัตถุดิบ','')}"
        )
        with st.expander(title):
            tc1, tc2 = st.columns(2)
            with tc1:
                st.markdown(
                    "**🚚 การส่งมอบ**  \n"
                    f"- ผู้ส่งมอบ: `{row.get('ผู้ส่งมอบ','')}`  \n"
                    f"- วันที่ส่ง: `{row.get('วันที่ส่งวัตถุดิบ','')}` เวลา `{row.get('เวลาส่ง','')}`  \n"
                    f"- บันทึกเมื่อ: `{row.get('วันที่บันทึก','')}`  \n"
                    f"- ผู้กรอก: `{row.get('ลงชื่อผู้กรอก','')}`  \n"
                    f"- อีเมล: `{row.get('อีเมล','')}`"
                )
                st.markdown(
                    "**🥬 วัตถุดิบ**  \n"
                    f"- ชนิด: `{row.get('ชนิดวัตถุดิบ','')}`  \n"
                    f"- Code: `{row.get('Code','')}`  \n"
                    f"- จำนวน: `{row.get('จำนวน(KG)','')}` KG  \n"
                    f"- สายพันธุ์: `{row.get('สายพันธุ์','')}`  \n"
                    f"- การปลูก: `{row.get('ลักษณะการปลูก','')}` / `{row.get('ลักษณะสถานที่ปลูก','')}`"
                )
            with tc2:
                st.markdown(
                    "**🌱 เก็บเกี่ยว / ล้าง**  \n"
                    f"- เก็บเกี่ยว: `{row.get('วันที่เก็บเกี่ยว','')}` เวลา `{row.get('เวลาเก็บเกี่ยว','')}`  \n"
                    f"- ล้างความสะอาด: `{row.get('วันที่ล้างความสะอาด','')}` เวลา `{row.get('เวลาล้างความสะอาด','')}`"
                )
                addr = (
                    f"{row.get('ที่อยู่เลขที่','')} "
                    f"หมู่ {row.get('หมู่ที่','')} "
                    f"ต.{row.get('ตำบล/เขต','')} "
                    f"อ.{row.get('อำเภอ/เมือง','')} "
                    f"จ.{row.get('จังหวัด','')} {row.get('รหัสไปรษณีย์','')}"
                )
                st.markdown(
                    "**🌾 แหล่งปลูก**  \n"
                    f"- ผู้ปลูก: `{row.get('ชื่อผู้ปลูก','')}`  \n"
                    f"- เลขที่ GAP: `{row.get('เลขที่GAP','')}`  \n"
                    f"- รหัสไร่: `{row.get('รหัสไร่','')}`  \n"
                    f"- ที่อยู่: {addr}"
                )
            # Timeline แบบสั้น
            st.markdown("---")
            st.markdown(
                f"**🧭 Timeline:** "
                f"`{row.get('วันที่เก็บเกี่ยว','—')}` (เก็บเกี่ยว) → "
                f"`{row.get('วันที่ล้างความสะอาด','—')}` (ล้าง) → "
                f"`{row.get('วันที่ส่งวัตถุดิบ','—')}` (ส่ง CPRAM) → "
                f"`{row.get('วันที่บันทึก','—')}` (บันทึกระบบ)"
            )
    if len(result) > 50:
        st.caption(f"⚠️ แสดงเฉพาะ 50 รายการแรก (จากทั้งหมด {len(result)} รายการ) — กรองเพิ่มเพื่อดูเฉพาะที่ต้องการ")

def render_data_management(df: pd.DataFrame):
    """ตารางข้อมูลดิบ + ดาวน์โหลด + ลบ"""
    if df.empty:
        st.info("ยังไม่มีข้อมูล")
    else:
        st.info(f"📊 ข้อมูลทั้งหมด: **{len(df)}** แถว")
        st.dataframe(df, use_container_width=True, height=320, hide_index=True)
    # ── ดาวน์โหลด ────────────────────────────────────────────────
    if os.path.exists(EXCEL_PATH):
        excel_bytes = get_encrypted_excel_bytes()
        if HAS_MSOFFCRYPTO:
            st.info(f"🔐 ไฟล์เข้ารหัสไว้แล้ว — รหัสเปิดไฟล์: **`{EXCEL_PASSWORD}`**")
        st.download_button(
            label="📊 ดาวน์โหลดไฟล์ Excel (ข้อมูลทั้งหมด)",
            data=excel_bytes,
            file_name="data_cpram_encrypted.xlsx" if HAS_MSOFFCRYPTO else "data_cpram.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    # ── ส่วนจัดการลบข้อมูล ───────────────────────────────────────
    st.markdown("---")
    st.markdown("### ⚠️ จัดการข้อมูล (โปรดใช้อย่างระมัดระวัง)")
    del_c1, del_c2 = st.columns(2)
    with del_c1:
        if not st.session_state.confirm_clear:
            if st.button("🧹 ลบข้อมูลทั้งหมด (เก็บหัวคอลัมน์)", use_container_width=True, key="btn_ask_clear"):
                st.session_state.confirm_clear = True
                st.rerun()
        else:
            st.warning("⚠️ ต้องการลบข้อมูลทั้งหมดจริงหรือไม่?  \nหัวคอลัมน์จะถูกเก็บไว้ แต่ข้อมูลทุกแถวจะหายไป")
            cc1, cc2 = st.columns(2)
            with cc1:
                if st.button("✅ ยืนยันลบ", use_container_width=True, key="btn_confirm_clear"):
                    ok, err = clear_excel_data()
                    st.session_state.confirm_clear = False
                    if ok: st.success("✅ ลบข้อมูลเรียบร้อย")
                    else:  st.error(f"❌ ลบไม่สำเร็จ: {err}")
                    st.rerun()
            with cc2:
                if st.button("❌ ยกเลิก", use_container_width=True, key="btn_cancel_clear"):
                    st.session_state.confirm_clear = False
                    st.rerun()
    with del_c2:
        if not st.session_state.confirm_delete_file:
            if st.button("🗑️ ลบไฟล์ Excel ทั้งหมด", use_container_width=True, key="btn_ask_del_file"):
                st.session_state.confirm_delete_file = True
                st.rerun()
        else:
            st.warning("⚠️ ต้องการลบไฟล์ Excel ทั้งหมดจริงหรือไม่?  \nไฟล์จะหายไปและจะถูกสร้างใหม่อัตโนมัติเมื่อมีข้อมูลส่งเข้ามาครั้งถัดไป")
            cd1, cd2 = st.columns(2)
            with cd1:
                if st.button("✅ ยืนยันลบไฟล์", use_container_width=True, key="btn_confirm_del_file"):
                    ok, err = delete_excel_file()
                    st.session_state.confirm_delete_file = False
                    if ok: st.success("✅ ลบไฟล์เรียบร้อย")
                    else:  st.error(f"❌ ลบไม่สำเร็จ: {err}")
                    st.rerun()
            with cd2:
                if st.button("❌ ยกเลิก", use_container_width=True, key="btn_cancel_del_file"):
                    st.session_state.confirm_delete_file = False
                    st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# 🔒 Admin panel (Dashboard / Traceability / Manage)
# ══════════════════════════════════════════════════════════════════════════════
with st.expander("🔒 สำหรับผู้ดูแลระบบ", expanded=False):
    admin_pw = st.text_input("รหัสผ่านผู้ดูแลระบบ", type="password", key="admin_pw_input")
    if admin_pw == ADMIN_PASSWORD:
        st.success("✅ เข้าสู่ระบบสำเร็จ")
        if not HAS_MSOFFCRYPTO:
            st.warning("⚠️ ยังไม่ได้ติดตั้ง `msoffcrypto-tool` — ไฟล์ที่ดาวน์โหลดจะไม่ได้เข้ารหัส\n\nรัน: `pip install msoffcrypto-tool`")
        # ── โหลดข้อมูล Excel (ถ้ามี) ─────────────────────────────────
        if os.path.exists(EXCEL_PATH):
            try:
                df_all = pd.read_excel(EXCEL_PATH)
            except Exception as ex:
                st.error(f"อ่านไฟล์ Excel ไม่สำเร็จ: {ex}")
                df_all = pd.DataFrame()
        else:
            df_all = pd.DataFrame()
            st.caption("ยังไม่มีข้อมูล Excel — รอให้มีการส่งข้อมูลเข้ามาก่อน")
        # ── Tabs ────────────────────────────────────────────────────
        tab_dash, tab_trace, tab_data = st.tabs([
            "📊 Dashboard", "🔍 Traceability", "📑 ข้อมูลดิบ / จัดการ"
        ])
        with tab_dash:
            render_dashboard(df_all)
        with tab_trace:
            render_traceability(df_all)
        with tab_data:
            render_data_management(df_all)
    elif admin_pw:
        st.error("❌ รหัสผ่านไม่ถูกต้อง")
    else:
        st.caption("กรุณากรอกรหัสผ่านเพื่อเข้าถึงข้อมูล")
        
